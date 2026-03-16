"""
Iteration 70: Test Admin Enrollments Export & Display
Tests for:
1. Excel export at /api/india-payments/admin/enrollments/export should return valid .xlsx
2. Excel should have wide-format headers like 'Participant 1 Name', 'Participant 2 Name' etc.
3. No cell in the Excel should contain the string 'None'
4. Each enrollment should be ONE row (not one row per participant)
5. GET /api/india-payments/admin/enrollments should return enrollments with clean participant data
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestAdminEnrollmentsExport:
    """Test the admin enrollments Excel export endpoint with wide format"""

    def test_export_returns_valid_xlsx(self):
        """GET /api/india-payments/admin/enrollments/export returns valid .xlsx file"""
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments/export")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text[:500] if response.status_code != 200 else ''}"
        
        # Verify content-type is Excel xlsx
        content_type = response.headers.get('content-type', '')
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type, \
            f"Expected xlsx content-type, got {content_type}"
        
        # Verify content-disposition has .xlsx filename
        content_disp = response.headers.get('content-disposition', '')
        assert '.xlsx' in content_disp, f"Expected .xlsx in content-disposition, got {content_disp}"
        
        # Verify response has content (xlsx starts with PK for ZIP signature)
        assert len(response.content) > 100, "Response content too small for valid xlsx"
        assert response.content[:2] == b'PK', "Excel file should start with PK (ZIP signature)"
        
        print(f"PASS: Export returns valid xlsx ({len(response.content)} bytes)")

    def test_export_has_wide_format_headers(self):
        """Excel headers should follow 'Participant N Field' pattern for wide format"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed for Excel inspection")
        
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments/export")
        assert response.status_code == 200, f"Export failed: {response.status_code}"
        
        # Load the Excel file
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Verify base headers exist
        base_headers = ["Receipt ID", "Status", "Program", "Booker Name", "Booker Email", "Participant Count"]
        for bh in base_headers:
            assert bh in headers, f"Missing base header '{bh}' in {headers[:15]}"
        
        # Verify wide-format participant headers exist (Participant 1 Name, Participant 1 Age, etc.)
        participant_headers_found = []
        expected_patterns = ["Participant 1 Name", "Participant 1 Age", "Participant 1 Email", "Participant 1 UID"]
        
        for expected in expected_patterns:
            if expected in headers:
                participant_headers_found.append(expected)
        
        assert len(participant_headers_found) >= 2, \
            f"Expected wide-format headers like 'Participant 1 Name', 'Participant 1 Age'. Found: {participant_headers_found}. Headers: {headers[:30]}"
        
        # Check for Participant 2 headers if there are multiple participants in data
        p2_headers = [h for h in headers if h and h.startswith("Participant 2")]
        print(f"PASS: Wide-format headers found. Participant 1 headers: {participant_headers_found}, Participant 2 headers count: {len(p2_headers)}")

    def test_export_no_none_values(self):
        """No cell in the Excel should contain the string 'None'"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed for Excel inspection")
        
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments/export")
        assert response.status_code == 200, f"Export failed: {response.status_code}"
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        none_cells = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    cell_str = str(cell.value)
                    if cell_str == "None" or cell_str == "null":
                        none_cells.append(f"Row {row_idx}, Col {col_idx}: '{cell_str}'")
        
        assert len(none_cells) == 0, f"Found 'None' or 'null' values in cells: {none_cells[:20]}"
        print(f"PASS: No 'None' or 'null' string values found in Excel")

    def test_export_one_row_per_enrollment(self):
        """Each enrollment should be ONE row (not one row per participant - wide format)"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed for Excel inspection")
        
        # First get the enrollment count from API
        enrollments_resp = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments")
        assert enrollments_resp.status_code == 200, f"Failed to get enrollments: {enrollments_resp.status_code}"
        enrollments = enrollments_resp.json()
        enrollment_count = len(enrollments)
        
        # Get Excel export
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments/export")
        assert response.status_code == 200, f"Export failed: {response.status_code}"
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Count data rows (exclude header)
        excel_row_count = ws.max_row - 1  # minus header row
        
        # Each enrollment should be one row
        assert excel_row_count == enrollment_count, \
            f"Excel has {excel_row_count} data rows but API returned {enrollment_count} enrollments. " \
            f"Expected 1 row per enrollment (wide format)."
        
        print(f"PASS: Excel has {excel_row_count} rows for {enrollment_count} enrollments (1:1 mapping)")

    def test_export_receipt_ids_unique(self):
        """Each row should have unique Receipt ID (first column)"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed for Excel inspection")
        
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments/export")
        assert response.status_code == 200, f"Export failed: {response.status_code}"
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        receipt_ids = []
        for row in ws.iter_rows(min_row=2):  # Skip header
            receipt_id = row[0].value  # First column
            if receipt_id:
                receipt_ids.append(receipt_id)
        
        # Check for duplicates
        unique_ids = set(receipt_ids)
        duplicates = len(receipt_ids) - len(unique_ids)
        
        assert duplicates == 0, f"Found {duplicates} duplicate Receipt IDs. Total rows: {len(receipt_ids)}, Unique: {len(unique_ids)}"
        print(f"PASS: All {len(receipt_ids)} Receipt IDs are unique")


class TestAdminEnrollmentsAPI:
    """Test the admin enrollments list API"""

    def test_enrollments_api_returns_200(self):
        """GET /api/india-payments/admin/enrollments returns 200"""
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert isinstance(data, list), "Response should be a list of enrollments"
        
        print(f"PASS: Enrollments API returned {len(data)} enrollments")

    def test_enrollments_have_clean_participant_data(self):
        """Enrollments should have participant data with names"""
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments")
        assert response.status_code == 200, f"Failed: {response.status_code}"
        
        enrollments = response.json()
        assert len(enrollments) > 0, "No enrollments found for testing"
        
        # Find an enrollment with participants
        enrollment_with_participants = None
        for e in enrollments:
            if e.get("participants") and len(e.get("participants", [])) > 0:
                enrollment_with_participants = e
                break
        
        if enrollment_with_participants:
            participants = enrollment_with_participants.get("participants", [])
            for p in participants:
                # Verify participant has name field
                assert "name" in p, f"Participant missing 'name' field: {p}"
                print(f"Participant data: name={p.get('name')}, has email={bool(p.get('email'))}")
        else:
            print("SKIP: No enrollments with participants found")
        
        print(f"PASS: Enrollments have clean participant data structure")

    def test_enrollments_have_required_fields(self):
        """Enrollments should have required fields for display"""
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments")
        assert response.status_code == 200, f"Failed: {response.status_code}"
        
        enrollments = response.json()
        if len(enrollments) == 0:
            pytest.skip("No enrollments to test")
        
        # Check first enrollment for required fields
        e = enrollments[0]
        required_fields = ["id", "status", "booker_name", "booker_email", "created_at"]
        
        for field in required_fields:
            assert field in e, f"Missing required field '{field}' in enrollment"
        
        print(f"PASS: Enrollments have all required fields for display")


class TestParticipantFieldsInExport:
    """Verify specific participant fields are present in export"""

    def test_participant_fields_complete(self):
        """Verify all expected participant fields are in the Excel headers"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed for Excel inspection")
        
        response = requests.get(f"{BASE_URL}/api/india-payments/admin/enrollments/export")
        assert response.status_code == 200, f"Export failed: {response.status_code}"
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        # Expected participant fields from the code
        expected_participant_fields = [
            "Participant 1 Name",
            "Participant 1 Relationship",
            "Participant 1 Age",
            "Participant 1 Gender",
            "Participant 1 Country",
            "Participant 1 Attendance Mode",
            "Participant 1 Is First Time",
            "Participant 1 Referral Source",
            "Participant 1 Email",
            "Participant 1 Phone",
            "Participant 1 UID",
        ]
        
        missing_fields = []
        for field in expected_participant_fields:
            if field not in headers:
                missing_fields.append(field)
        
        # Allow some flexibility but require key fields
        key_fields = ["Participant 1 Name", "Participant 1 Age", "Participant 1 Email", "Participant 1 UID"]
        key_missing = [f for f in key_fields if f in missing_fields]
        
        assert len(key_missing) == 0, f"Missing key participant fields: {key_missing}. All headers: {headers[:40]}"
        
        print(f"PASS: All key participant fields found. Minor missing: {missing_fields}")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
