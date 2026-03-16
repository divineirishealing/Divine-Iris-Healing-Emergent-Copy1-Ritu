from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from typing import Optional
import os, uuid, shutil, logging
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path
from datetime import datetime, timezone

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

router = APIRouter(prefix="/api/india-payments", tags=["India Payments"])

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

logger = logging.getLogger(__name__)

UPLOAD_DIR = ROOT_DIR / "uploads" / "payment_proofs"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


@router.post("/submit-proof")
async def submit_payment_proof(
    enrollment_id: str = Form(...),
    payer_name: str = Form(...),
    payer_email: str = Form(""),
    payer_phone: str = Form(""),
    payment_date: str = Form(...),
    bank_name: str = Form(""),
    transaction_id: str = Form(...),
    amount: str = Form(...),
    city: str = Form(""),
    state: str = Form(""),
    payment_method: str = Form(""),
    program_type: str = Form(""),
    selected_item: str = Form(""),
    is_emi: str = Form("false"),
    emi_total_months: str = Form(""),
    emi_months_covered: str = Form(""),
    screenshot: UploadFile = File(...),
    notes: str = Form(""),
):
    """Submit India alternative payment proof for admin approval."""
    # Validate enrollment exists (skip for standalone 'MANUAL' submissions)
    enrollment = {}
    if enrollment_id != "MANUAL":
        enrollment = await db.enrollments.find_one({"id": enrollment_id}, {"_id": 0}) or {}

    # Save screenshot
    ext = screenshot.filename.split(".")[-1] if "." in screenshot.filename else "png"
    filename = f"{uuid.uuid4().hex[:12]}.{ext}"
    filepath = UPLOAD_DIR / filename
    with open(filepath, "wb") as f:
        shutil.copyfileobj(screenshot.file, f)

    proof = {
        "id": str(uuid.uuid4()),
        "enrollment_id": enrollment_id,
        "booker_name": enrollment.get("booker_name", payer_name),
        "booker_email": enrollment.get("booker_email", payer_email),
        "payer_name": payer_name,
        "payer_email": payer_email,
        "payer_phone": payer_phone,
        "payment_date": payment_date,
        "bank_name": bank_name,
        "transaction_id": transaction_id,
        "program_type": program_type,
        "selected_item": selected_item,
        "program_title": enrollment.get("item_title", selected_item or program_type),
        "amount": amount,
        "city": city,
        "state": state,
        "payment_method": payment_method,
        "is_emi": is_emi == "true",
        "emi_total_months": int(emi_total_months) if emi_total_months else None,
        "emi_months_covered": int(emi_months_covered) if emi_months_covered else None,
        "notes": notes,
        "screenshot_url": f"/api/uploads/payment_proofs/{filename}",
        "status": "pending",
        "participants": enrollment.get("participants", []),
        "participant_count": enrollment.get("participant_count", 1),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.india_payment_proofs.insert_one(proof)

    # Update enrollment status if linked
    if enrollment_id != "MANUAL" and enrollment:
        await db.enrollments.update_one(
            {"id": enrollment_id},
            {"$set": {
                "status": "india_payment_proof_submitted",
                "india_payment_proof_id": proof["id"],
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )

    logger.info(f"[INDIA PAYMENT PROOF] enrollment={enrollment_id}, txn={transaction_id}, amount={amount}")
    return {"message": "Payment proof submitted successfully. Awaiting admin approval.", "proof_id": proof["id"]}


@router.get("/admin/list")
async def list_payment_proofs():
    """Admin: list all India payment proofs."""
    proofs = await db.india_payment_proofs.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return proofs


@router.post("/admin/{proof_id}/approve")
async def approve_payment_proof(proof_id: str):
    """Admin: approve India payment proof and complete enrollment."""
    proof = await db.india_payment_proofs.find_one({"id": proof_id}, {"_id": 0})
    if not proof:
        raise HTTPException(status_code=404, detail="Payment proof not found")

    # Mark proof as approved
    await db.india_payment_proofs.update_one(
        {"id": proof_id},
        {"$set": {"status": "approved", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    enrollment_id = proof.get("enrollment_id")
    if enrollment_id:
        # Create a completed transaction
        fake_session_id = f"india_{uuid.uuid4().hex[:12]}"
        transaction = {
            "id": str(uuid.uuid4()),
            "enrollment_id": enrollment_id,
            "stripe_session_id": fake_session_id,
            "item_type": "program",
            "item_id": "",
            "item_title": proof.get("program_title", ""),
            "amount": float(proof.get("amount", 0)),
            "currency": "inr",
            "payment_status": "paid",
            "booker_name": proof.get("booker_name"),
            "booker_email": proof.get("booker_email"),
            "participants": proof.get("participants", []),
            "participant_count": proof.get("participant_count", 1),
            "is_india_alt": True,
            "india_proof_id": proof_id,
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
        }
        await db.payment_transactions.insert_one(transaction)

        # Complete enrollment
        await db.enrollments.update_one(
            {"id": enrollment_id},
            {"$set": {
                "step": 5,
                "status": "completed",
                "stripe_session_id": fake_session_id,
                "is_india_alt": True,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )

        # Generate UIDs and send emails
        try:
            from routes.payments import generate_participant_uids, send_enrollment_emails
            await generate_participant_uids(fake_session_id)
            import asyncio
            asyncio.create_task(send_enrollment_emails(fake_session_id))
        except Exception as e:
            logger.warning(f"Error generating UIDs/emails for India payment: {e}")

    return {"message": "Payment proof approved. Enrollment completed.", "status": "approved"}


@router.post("/admin/{proof_id}/reject")
async def reject_payment_proof(proof_id: str, reason: str = ""):
    """Admin: reject India payment proof."""
    proof = await db.india_payment_proofs.find_one({"id": proof_id})
    if not proof:
        raise HTTPException(status_code=404, detail="Payment proof not found")

    await db.india_payment_proofs.update_one(
        {"id": proof_id},
        {"$set": {"status": "rejected", "reject_reason": reason, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    enrollment_id = proof.get("enrollment_id")


@router.get("/admin/enrollments")
async def list_enrollments():
    """Admin: list all enrollments."""
    enrollments = await db.enrollments.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    # Also fetch associated transactions
    for e in enrollments:
        txn = await db.payment_transactions.find_one(
            {"enrollment_id": e.get("id")}, {"_id": 0, "amount": 1, "currency": 1, "payment_status": 1}
        )
        e["payment"] = txn if txn else None
    return enrollments


@router.get("/admin/enrollments/export")
async def export_enrollments_excel():
    """Admin: export all enrollments as Excel — wide format, one row per enrollment, database-ready."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    def clean(val):
        """Convert any value to a clean string. None/null → empty string."""
        if val is None:
            return ""
        s = str(val)
        if s in ("None", "null"):
            return ""
        return s

    enrollments = await db.enrollments.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)

    # Determine max participant count across all enrollments
    max_participants = 0
    for e in enrollments:
        count = len(e.get("participants") or [])
        if count > max_participants:
            max_participants = count
    max_participants = max(max_participants, 1)

    # Build headers: base columns + per-participant columns
    base_headers = [
        "Receipt ID", "Status", "Program", "Program Type",
        "Booker Name", "Booker Email", "Booker Country", "Booker Phone",
        "Participant Count", "Enrollment Date",
    ]

    participant_fields = [
        "Name", "Relationship", "Age", "Gender", "Country",
        "Attendance Mode", "Is First Time", "Referral Source", "Referred By",
        "Email", "Phone", "WhatsApp", "UID",
    ]

    headers = list(base_headers)
    for i in range(1, max_participants + 1):
        for field in participant_fields:
            headers.append(f"Participant {i} {field}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrollments"

    # Style header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Build column letter ref for auto_filter
    last_col = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes = "A2"

    for e in enrollments:
        # Format created_at
        created_at = clean(e.get("created_at"))
        if created_at:
            try:
                from datetime import datetime as dt
                d = dt.fromisoformat(created_at.replace("Z", "+00:00")) if isinstance(created_at, str) else created_at
                created_at = d.strftime("%Y-%m-%d %H:%M")
            except Exception:
                pass

        # Base row data
        row = [
            clean(e.get("id")),
            clean(e.get("status")),
            clean(e.get("item_title")),
            clean(e.get("item_type")),
            clean(e.get("booker_name")),
            clean(e.get("booker_email")),
            clean(e.get("booker_country")),
            clean(e.get("phone")),
            str(e.get("participant_count", 0) or 0),
            created_at,
        ]

        # Append participant columns (wide format)
        participants = e.get("participants") or []
        for i in range(max_participants):
            if i < len(participants):
                p = participants[i]
                p_phone = clean(p.get("phone"))
                p_wa = clean(p.get("whatsapp"))
                row.extend([
                    clean(p.get("name")),
                    clean(p.get("relationship")),
                    clean(p.get("age")),
                    clean(p.get("gender")),
                    clean(p.get("country")),
                    clean(p.get("attendance_mode")),
                    "Yes" if p.get("is_first_time") else "No",
                    clean(p.get("referral_source")),
                    clean(p.get("referred_by_name")),
                    clean(p.get("email")),
                    p_phone,
                    p_wa,
                    clean(p.get("uid")),
                ])
            else:
                # Empty columns for missing participants
                row.extend([""] * len(participant_fields))

        ws.append(row)

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enrollments.xlsx"}
    )
